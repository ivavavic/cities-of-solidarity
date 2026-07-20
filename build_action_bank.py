"""Generate data/action_bank.csv and data/action_bank.xlsx.
Sources: the 12 Comrat LAP priority actions (hand-encoded below) and the
simplified example actions parsed from LAP_Example_Action_Bank.docx."""
import csv, re
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

COLS = ["action_id","action_title","sector","city","short_description","long_description",
        "target_group","implementation_owner","partners","indicative_cost","timeline",
        "status","scale_up_potential","tags","source_lap","display_order"]

rows = []

# -------------------- 12 Comrat LAP priority actions --------------------
comrat = [
 dict(t="Municipal Refugee Focal Point", s="Governance & Coordination",
   sd="A formally designated focal point in City Hall serving as the single entry point for refugee-related matters, referrals and complex cases.",
   ld="Designate a municipal Refugee Focal Point who understands the legal framework, rights and procedures applicable to refugees and Temporary Protection holders; maintains up-to-date information on needs; makes referrals across departments and national institutions; supports complex cases; and anchors municipal leadership in the Local Refugee Coordination Forum.",
   tg="Refugees, TP holders and municipal departments", own="Comrat City Hall",
   pa="UNHCR; LRCF partners; MLSP", cost="Low (staff time + training)", tl="2026 — quick start",
   st="investment-ready", su="High — replicable in any municipality at low cost", tags="governance;referrals;coordination"),
 dict(t="Integrated Community Care", s="Social Protection & Care",
   sd="Home-based care, facility upgrades and a trained buddy network for older residents, persons with disabilities and isolated households.",
   ld="Upgrade local care by linking facility improvements with home-based support and trained volunteers. Social workers, nurses and community buddies provide home visits, well-being checks, medication support, food access and companionship. Responds to Comrat's ageing caseload: 27.9% of surveyed members are aged 60+ and 45.6% of households report a chronic condition.",
   tg="Older persons, persons with disabilities, single caregivers", own="Comrat Social Assistance Directorate",
   pa="HelpAge; NGOs; volunteers; UNHCR", cost="Medium", tl="2026–2027, phased",
   st="investment-ready", su="High — model transferable to other ageing districts", tags="care;health;home visits;older persons"),
 dict(t="Telemedicine & Health Workforce Housing", s="Health",
   sd="Service housing to attract and retain medical staff, paired with a telemedicine point connecting patients to specialists.",
   ld="Strengthen local healthcare capacity by attracting doctors, nurses and specialists — including qualified refugee professionals where legally possible — through service housing and recruitment support, while equipping a telemedicine point in the polyclinic to connect complex cases to specialists in larger cities and reduce unnecessary travel.",
   tg="All residents; patients with complex or chronic conditions", own="Comrat City Hall + district health authorities",
   pa="Ministry of Health; development partners", cost="Medium–high (works + equipment)", tl="2026–2028",
   st="concept", su="Medium — depends on national health workforce policy", tags="health;telemedicine;workforce;housing"),
 dict(t="Talent & Opportunity Fund", s="Education & Childcare",
   sd="Scholarships and stipends for refugee and vulnerable local students, linked to service-learning such as tutoring younger pupils.",
   ld="Partner with Comrat State University, colleges and vocational institutions to provide scholarships or stipend top-ups combining academic potential with vulnerability. Recipients commit to service-learning — tutoring, learning-hub support — creating a self-sustaining loop between higher education and community learning support.",
   tg="Refugee and vulnerable local youth and students", own="Comrat City Hall + Comrat State University",
   pa="Universities; colleges; donors; UNHCR", cost="Low–medium (fund size dependent)", tl="Academic year 2026/27 onward",
   st="pilot", su="High — scalable fund model; replicable per city", tags="education;scholarships;youth;service learning"),
 dict(t="Public Space Revitalization", s="Housing & Infrastructure",
   sd="Upgrades to underused public spaces and civic infrastructure serving refugee and host neighbourhoods alike.",
   ld="Revitalize underused urban land, public spaces and civic infrastructure — lighting, green space, safe routes, seating, play areas — in neighbourhoods where needs concentrate. Investments upgrade shared spaces used by all residents rather than creating separate refugee facilities, and can incorporate energy-efficient and nature-based solutions.",
   tg="All residents of targeted neighbourhoods", own="Comrat City Hall (urban planning)",
   pa="Development partners; contractors; community groups", cost="Medium–high", tl="2026–2028, phased by site",
   st="investment-ready", su="High — bankable as part of regional public-works packages", tags="infrastructure;public space;urban"),
 dict(t="Social Mobility Service", s="Mobility",
   sd="A social taxi and community shuttle providing accessible, low-cost mobility for older persons, persons with disabilities and students.",
   ld="Establish a targeted municipal mobility service complementing regular transport: an accessible social-taxi vehicle for on-demand medical and social trips, plus a scheduled community shuttle linking peripheral neighbourhoods to schools, clinics and services. Provides universal mobility for refugees and local residents, with simple eligibility linked to social assistance.",
   tg="Older persons, persons with disabilities, students, low-income households", own="Comrat City Hall + Social Assistance",
   pa="NGOs; transport operators; donors", cost="Medium (vehicle + operations)", tl="2026 pilot, 2027 scale",
   st="pilot", su="Medium–high — operating model replicable across raions", tags="mobility;social taxi;accessibility"),
 dict(t="Municipal Social Rental Agency", s="Housing & Infrastructure",
   sd="A municipal mechanism connecting vulnerable tenants with private landlords: unit registry, lease formalization, mediation and a small guarantee fund.",
   ld="Create a Social Rental Agency that recruits property owners, maintains a registry of affordable units, formalizes leases and mediates disputes, with a small guarantee fund covering deposits and minor repairs. Responds to widespread informal tenure (57.5% of surveyed households lack a formal lease) and makes the private rental market safer for tenants and landlords alike.",
   tg="Refugee and vulnerable local tenants; private landlords", own="Comrat City Hall",
   pa="UNHCR; NGOs; landlord associations", cost="Medium", tl="2026–2027",
   st="investment-ready", su="High — core model for inter-municipal housing pipeline", tags="housing;rental;tenure;formalization"),
 dict(t="Agri-Social Hub", s="Food Security & Livelihoods",
   sd="A community garden and greenhouse, community kitchen and periodic eco-market combining food security, skills and social cohesion.",
   ld="Develop an agri-food and social inclusion hub where refugees and local residents grow produce, prepare meals, support a social canteen function and sell selected products through a periodic eco-market. Links food security, practical skills, local producer visibility and everyday interaction in one package, and can enhance climate resilience along the Yalpug river corridor.",
   tg="Vulnerable households; small producers; community members", own="Comrat City Hall + partner NGO",
   pa="NGOs; agricultural college; donors", cost="Medium", tl="2027, seasonal phasing",
   st="concept", su="Medium — components replicable; site-specific design", tags="food security;livelihoods;community;green"),
 dict(t="Mixed-Income PPP Housing", s="Housing & Infrastructure",
   sd="Public-private partnership housing on municipal land, reserving a share of units for vulnerable local and refugee families.",
   ld="Develop mixed-income affordable housing through partnership with private developers and development finance: the municipality contributes land or an unused asset and infrastructure connections; private partners finance construction. A share of units is reserved for vulnerable households while other units cross-subsidize operations, bundling energy efficiency and maintenance obligations into the PPP contract.",
   tg="Vulnerable local and refugee families; municipal housing stock", own="Comrat City Hall",
   pa="Private developers; IFIs; national authorities", cost="High (blended finance)", tl="2027–2029, preparation first",
   st="concept", su="High — flagship for aggregated IFI investment across cities", tags="housing;PPP;investment;energy efficiency"),
 dict(t="Green Public Works", s="Livelihoods & Employment",
   sd="Short-term paid public works — greening, river corridor cleanup, public-space upkeep — with fair wages and skills certificates.",
   ld="Create short-term paid public-works and municipal internship opportunities for refugees and vulnerable locals: park maintenance, tree planting, cleanup of the Yalpug corridor, waste-separation pilots and public-space upkeep. Participants receive fair wages, safety training, certificates and referrals to longer-term employment, while the city gains visible environmental improvements.",
   tg="Unemployed refugees and vulnerable local residents", own="Comrat City Hall (municipal services)",
   pa="Employment agency; NGOs; donors", cost="Low–medium per cohort", tl="2026 onward, seasonal cohorts",
   st="pilot", su="High — cohort model easily repeated and scaled", tags="employment;green jobs;public works"),
 dict(t="Women's Entrepreneurship Hub", s="Women's Economic Empowerment",
   sd="Business training, mentoring, micro-grants and childcare-friendly formats helping refugee and local women start or grow micro-businesses.",
   ld="Support refugee and local women to start or expand micro-businesses and social enterprises through training, mentoring, legal and start-up advice, micro-grants, buyer days and referrals to national SME instruments. Childcare-friendly formats respond directly to the caregiving barrier — around 25% of respondents cite care responsibilities as their main obstacle to work.",
   tg="Refugee and local women entrepreneurs and jobseekers", own="Comrat City Hall + business incubator",
   pa="ODIMM/national SME programmes; NGOs; UNHCR", cost="Medium", tl="2026–2028",
   st="investment-ready", su="High — hub model prepared for multi-city rollout", tags="women;entrepreneurship;livelihoods;childcare"),
 dict(t="Digital & AI Inclusion Hub", s="Digital Inclusion & Skills",
   sd="Practical digital and AI skills workshops — e-services, remote work, AI tools — hosted in the library and community spaces.",
   ld="Offer practical workshops helping refugees and residents use digital tools for work, study and services: e-government platforms, digital payments, job platforms, online safety, and AI tools for translation, CV drafting and productivity. Dedicated tracks for women, youth and older persons prepare participants for freelance and higher-skilled work and help diversify the local economy.",
   tg="Women, youth, older persons, jobseekers", own="Comrat City Hall + municipal library",
   pa="ICT partners; NGOs; telecom providers", cost="Low–medium", tl="2026 onward, rolling",
   st="pilot", su="High — curriculum replicable across all CoS cities", tags="digital;AI;skills;inclusion"),
]
for i, a in enumerate(comrat, 1):
    rows.append([f"COM-{i:02d}", a["t"], a["s"], "Comrat", a["sd"], a["ld"], a["tg"], a["own"],
                 a["pa"], a["cost"], a["tl"], a["st"], a["su"], a["tags"],
                 "Comrat Local Action Plan for Refugee Inclusion (2026)", i])

# -------------------- Bălți workshop example actions (parsed from docx) --------------------
SECTOR_MAP = {
 "Governance":"Governance & Coordination", "Accountability":"Governance & Coordination",
 "Integration":"Access to Services & Information",
 "Social Protection":"Social Protection & Care",
 "Healthcare":"Health",
 "Mobility":"Mobility",
 "Protection":"Protection",
 "Food Security":"Food Security & Livelihoods",
 "Employment":"Livelihoods & Employment",
 "Education":"Education & Childcare", "Childcare":"Education & Childcare",
 "Early Childhood Education":"Education & Childcare",
 "Digital Inclusion":"Digital Inclusion & Skills", "Digital Skills":"Digital Inclusion & Skills",
 "Housing":"Housing & Infrastructure", "Urban Planning":"Housing & Infrastructure",
 "Social Inclusion":"Social Cohesion", "Social Cohesion":"Social Cohesion",
}
def clean(s):
    return re.sub(r"\s+", " ", s.replace("\u00a0"," ")).strip()

doc = Document("/mnt/user-data/uploads/LAP_Example_Action_Bank.docx")
n = 0
for table in doc.tables:
    for r in table.rows[1:]:                         # skip header row
        title, sector_raw, desc = [clean(c.text) for c in r.cells[:3]]
        if not title or title.lower().startswith("project name"):
            continue
        n += 1
        parts = [clean(p) for p in sector_raw.split("/")]
        sector = SECTOR_MAP.get(parts[0], parts[0])
        first_sentence = desc.split(". ")[0].rstrip(".") + "."
        tags = ";".join(p.lower() for p in parts)
        rows.append([f"EX-{n:02d}", title, sector, "Adaptable (all cities)", first_sentence, desc,
                     "", "Municipality + partners (to be defined)", "", "To be costed locally",
                     "Flexible", "concept", "Adaptable to any municipality", tags,
                     "Bălți LAP workshop — example action bank", 100 + n])

# ---------- write CSV ----------
with open("data/action_bank.csv","w",newline="",encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(COLS); w.writerows(rows)

# ---------- write XLSX master ----------
wb = Workbook(); ws = wb.active; ws.title = "action_bank"
ws.append(COLS)
for cell in ws[1]:
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="0072BC")
for r in rows: ws.append(r)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=cell.column_letter in ("E","F"))
widths = dict(A=9,B=38,C=26,D=22,E=60,F=80,G=34,H=32,I=32,J=24,K=20,L=16,M=40,N=32,O=40,P=13)
for col,wd in widths.items(): ws.column_dimensions[col].width = wd
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

notes = wb.create_sheet("READ ME")
info = [
 ["How to update the Action Bank"],
 [""],
 ["1. Add or edit rows on the 'action_bank' sheet. Keep the column names in row 1 unchanged."],
 ["2. Save this Excel file (it is the human-editable master)."],
 ["3. Export the sheet as CSV: File > Save As > CSV UTF-8, and overwrite data/action_bank.csv."],
 ["   The website reads the CSV, not this Excel file."],
 [""],
 ["Field guidance:"],
 ["  action_id       Unique ID, e.g. COM-01 (Comrat), BAL-01 (Bălți), EX-01 (example)."],
 ["  sector          Drives the sector filter and card colour. Reuse existing sector names where possible."],
 ["  city            City name, or 'Adaptable (all cities)' for generic example actions."],
 ["  status          One of: concept | pilot | investment-ready | scalable."],
 ["  tags            Semicolon-separated keywords used by the search box."],
 ["  display_order   Lower numbers appear first."],
 ["  Leave any field blank if unknown — the website shows a clean placeholder."],
]
for line in info: notes.append(line)
notes["A1"].font = Font(name="Arial", bold=True, size=12)
for row in notes.iter_rows(min_row=2):
    for cell in row: cell.font = Font(name="Arial", size=10)
notes.column_dimensions["A"].width = 110
wb.save("data/action_bank.xlsx")
print(f"OK: {len(rows)} actions ({n} examples)")
