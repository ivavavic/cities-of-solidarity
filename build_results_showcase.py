"""Regenerate data/results_showcase.csv from data/results_showcase.xlsx.

Usage (from the repository root):
    python build_results_showcase.py

The website reads the CSV at runtime; the Excel file is the human-editable
master. Run this after editing the Excel file, then commit and push both
files. Requires the openpyxl package:  pip install openpyxl
"""
import csv
import datetime as dt
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is not installed. Run:  pip install openpyxl")

XLSX = "data/results_showcase.xlsx"
CSV_OUT = "data/results_showcase.csv"
SHEET = "Results Showcase"

# CSV columns, in order, matched to the Excel columns A..T
COLS = ["project_id", "project_name", "district", "municipality", "address",
        "latitude", "longitude", "facility_type", "ownership", "scope_of_works",
        "impact", "implementation_modality", "total_investment_usd",
        "funding_source", "status", "start_date", "completion_date",
        "before_photo", "after_photo", "remarks"]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

STATUS_MAP = {"ongoing": "In progress", "in progress": "In progress",
              "completed": "Completed", "planned": "Planned"}


def clean(value, col):
    """Normalise a cell value for the CSV."""
    if value is None:
        return ""
    # Dates typed as real Excel dates -> "Mar 2024"
    if isinstance(value, (dt.datetime, dt.date)):
        return f"{MONTHS[value.month - 1]} {value.year}"
    text = re.sub(r"\s+", " ", str(value)).strip()
    if col == "status":
        return STATUS_MAP.get(text.lower(), text)
    if col in ("latitude", "longitude"):
        try:
            return f"{float(text):.6f}"
        except ValueError:
            return ""
    if col == "total_investment_usd":
        try:
            return f"{float(text):.2f}"
        except ValueError:
            return ""
    if col in ("before_photo", "after_photo"):
        # keep real links (http(s)) and repo-relative paths (assets/...);
        # drop leftover SharePoint breadcrumb text
        low = text.lower()
        if low.startswith("http") or low.startswith("assets/"):
            return text
        return ""
    return text


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f'Sheet "{SHEET}" not found in {XLSX}')
    ws = wb[SHEET]

    rows_out = []
    for row in ws.iter_rows(min_row=2, max_col=len(COLS), values_only=True):
        record = {col: clean(val, col) for col, val in zip(COLS, row)}
        if not record["project_name"]:
            continue  # skip empty rows
        rows_out.append(record)

    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLS)
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out)} projects to {CSV_OUT}")


if __name__ == "__main__":
    main()
