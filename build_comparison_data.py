"""Generate data/city_comparison.csv and data/city_comparison.xlsx.
Run once during site creation; afterwards the Excel file is the human-editable
master and the CSV is what the website reads."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

COLS = ["city","indicator_group","indicator_name","value","display","unit","year",
        "source","note","severity","display_order"]

PCP = {
    "Comrat":   "REACH/IMPACT PCP Comrat, Apr 2025 (N=103 HH)",
    "Bălți":    "REACH/IMPACT PCP Bălți, Apr 2025 (N=105 HH)",
    "Chișinău": "REACH/IMPACT PCP Chișinău, Oct–Nov 2025 (N=532 HH)",
    "Cahul":    "REACH/IMPACT PCP Cahul, Jun 2026 (N=106 HH)",
}
YR = {"Comrat":2025,"Bălți":2025,"Chișinău":2025,"Cahul":2026}
ORDER = ["Comrat","Bălți","Chișinău","Cahul"]

rows = []
def add(city, group, name, value, display="", unit="%", year=None, source="", note="", severity="", order=None):
    if year is None: year = YR.get(city,"")
    if not source: source = PCP.get(city, "REACH/IMPACT PCP rounds 2025–2026")
    if order is None: order = len([r for r in rows if r[1]==group]) + 1
    rows.append([city, group, name, value, display, unit, year, source, note, severity, order])

# ---- headline stats (chapter 01) ----
add("All","headline","Refugee households surveyed",846,"","households","2025–2026","REACH/IMPACT PCP rounds, four cities","Coverage",order=1)
add("All","headline","Household members covered",2143,"","people","2025–2026","REACH/IMPACT PCP rounds, four cities","Coverage",order=2)
add("All","headline","Lack public health insurance or unsure of eligibility","","83–89%","%","2025–2026","REACH/IMPACT PCP rounds, four cities","Shared gap — range across the four cities",order=3)
add("All","headline","Out of work because of care responsibilities","","1 in 5+","","2025–2026","REACH/IMPACT PCP rounds, four cities","Shared barrier — in all four cities",order=4)

# ---- the big split (chapter 02) ----
for c,v in zip(ORDER,[11.7,19.0,30.1,40.6]):
    add(c,"employment","Employed or self-employed",v,note="% of respondents")
for c,v,n in zip(ORDER,[53.4,43.8,38.2,2.8],["","","","timing: surveyed after humanitarian cash phase-down"]):
    add(c,"aid_dependence","Humanitarian assistance as main household income source",v,note=n or "% of households")
for c,v,n in zip(ORDER,[13.6,22.9,27.8,59.4],["","","","timing: surveyed after humanitarian cash phase-down"]):
    add(c,"employment_income","Employment as main household income source",v,note=n or "% of households")

# ---- sample sizes ----
for c,n,p in [("Comrat",103,"Apr 2025"),("Bălți",105,"Apr 2025"),("Chișinău",532,"Oct–Nov 2025"),("Cahul",106,"Jun 2026")]:
    add(c,"sample","Households surveyed",n,"","households",YR[c],PCP[c],p)

# ---- archetype cards (chapter 03): 4 indicators per city ----
arch = {
 "Comrat":[("Members aged 60+",27.9),("Aid = main income",53.4),("Chronic condition in HH",45.6),("Hosted / housed free",51.4)],
 "Bălți":[("HHs with children",63.8),("Single adult + dependents",43.8),("No children enrolled*",35.7),("Unstable accommodation",41.9)],
 "Chișinău":[(">50% income on housing",66.7),("No tenure document",41.5),("Moved at least once",53.0),("Employed",30.1)],
 "Cahul":[("Employed",40.6),("Plan to stay",69.8),(">50% income on housing",59.4),("HHs with children",65.1)],
}
for c, items in arch.items():
    for i,(name,v) in enumerate(items,1):
        add(c,"archetype",name,v,order=i)

# ---- universal bottlenecks (chapter 04): severity 1 (worst) .. 4 (least) within row ----
bott = [
 ("No public health insurance",[("Comrat",83.5,"",2),("Bălți",88.6,"",1),("Chișinău",85.9,"",1),("Cahul",83.0,"",2)]),
 ("Out of work due to care duties",[("Comrat",25.2,"",1),("Bălți",22.9,"",2),("Chișinău",22.0,"",2),("Cahul",19.8,"",3)]),
 ("No Romanian proficiency",[("Comrat",63.1,"",1),("Bălți",59.0,"",2),("Chișinău",50.9,"",3),("Cahul",55.7,"",2)]),
 (">50% of income on housing & utilities",[("Comrat",35.9,"",4),("Bălți",56.2,"",2),("Chișinău",66.7,"",1),("Cahul",59.4,"",2)]),
 ("UA online curriculum preferred*",[("Comrat","","not published",""),("Bălți","","18 of 20",1),("Chișinău",79.7,"",1),("Cahul","","8 of 12",2)]),
]
for bi,(name,cells) in enumerate(bott,1):
    for c,v,disp,sev in cells:
        note = "Dominant reason for non-enrolment; counts shown where n<30" if "curriculum" in name else ""
        rows.append([c,"bottleneck",name,v,disp,"%",YR[c],PCP[c],note,sev,bi])

# ---- housing (chapter 05) ----
for c,v in zip(ORDER,[35.9,56.2,66.7,59.4]):
    add(c,"housing_burden","Spend >50% of income on housing & utilities",v)
inf = [("Chișinău","No tenure document†",41.5,""),("Chișinău","Moved ≥ once",53.0,""),
       ("Bălți","Unstable accommodation",41.9,""),("Cahul","Unstable accommodation",22.6,""),
       ("Comrat","Renters without documents†",35.0,"Renters only (n=40); a further 22.5% preferred not to answer")]
for i,(c,n,v,note) in enumerate(inf,1):
    add(c,"housing_informality",n,v,note=note,order=i)
ten = [("Comrat","Hosted / housed free",51.4),("Comrat","Renting",38.8),("Bălți","Renting",66.7),
       ("Cahul","Renting",70.7),("Comrat","≥1 housing problem",25.2),("Cahul","≥1 housing problem",22.6)]
for i,(c,n,v) in enumerate(ten,1):
    add(c,"housing_tenure",n,v,order=i)

# ---- work & care (chapter 06) ----
slope = [("Cahul",68.9,40.6),("Chișinău",59.0,30.1),("Bălți",66.7,19.0),("Comrat","",11.7)]
for c,before,now in slope:
    note = "Pre-displacement employment topline not published" if before=="" else ""
    add(c,"employment_slope","Employed before displacement",before,note=note,order=1)
    add(c,"employment_slope","Employed now in Moldova",now,order=2)
for c,v in zip(ORDER,[25.2,22.9,22.0,19.8]):
    add(c,"care_barrier","Not working due to care responsibilities",v)

# ---- services (chapter 07) ----
for c,v in zip(ORDER,[12.6,9.5,3.6,0.9]):
    add(c,"health_affordability","Could not afford healthcare",v)
for c,v in zip(ORDER,[66.7,64.3,75.8,76.9]):
    add(c,"education_enrolment","All school-aged children enrolled",v,note="% of HHs with school-aged children; sub-samples Comrat n=39, Bălți n=56, Chișinău n=244, Cahul n=52")
cc = [("Comrat",46.7,"7/15"),("Bălți",68.4,"13/19"),("Chișinău",62.0,"62%"),("Cahul",71.0,"71%")]
for c,v,d in cc:
    add(c,"childcare_access","Can access childcare when needed",v,display=d,note="Small sub-samples (n=15–92); counts shown where n<30")
for c,v in zip(ORDER,[39.8,41.0,49.2,84.9]):
    add(c,"information","No information needs reported",v,note="Digital-confidence module measured only in Chișinău and Cahul")

# ---------- write CSV ----------
with open("data/city_comparison.csv","w",newline="",encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(COLS)
    w.writerows(rows)

# ---------- write XLSX master ----------
wb = Workbook(); ws = wb.active; ws.title = "city_comparison"
head_fill = PatternFill("solid", start_color="0072BC")
ws.append(COLS)
for cell in ws[1]:
    cell.font = Font(name="Arial", bold=True, color="FFFFFF"); cell.fill = head_fill
    cell.alignment = Alignment(vertical="center")
for r in rows:
    ws.append(r)
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.font = Font(name="Arial", size=10)
widths = dict(A=12,B=20,C=46,D=9,E=12,F=10,G=11,H=44,I=52,J=9,K=13)
for col,wd in widths.items(): ws.column_dimensions[col].width = wd
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

notes = wb.create_sheet("READ ME")
info = [
 ["How to update the Four Cities comparison data"],
 [""],
 ["1. Edit values on the 'city_comparison' sheet. Keep the column names in row 1 unchanged."],
 ["2. Save this Excel file (it is the human-editable master)."],
 ["3. Export the sheet as CSV: File > Save As > CSV UTF-8, and overwrite data/city_comparison.csv."],
 ["   The website reads the CSV, not this Excel file."],
 [""],
 ["Column meanings:"],
 ["  city            City name (Comrat, Bălți, Chișinău, Cahul) or 'All' for cross-city figures."],
 ["  indicator_group Which chart on comparison.html uses the row (do not rename existing groups)."],
 ["  indicator_name  Label shown next to the value."],
 ["  value           Numeric value (percent unless the unit says otherwise). May be blank if 'display' is set."],
 ["  display         Optional text shown instead of the numeric value (e.g. '18 of 20', '83–89%')."],
 ["  unit            %, households, people …"],
 ["  year            Survey year."],
 ["  source          Data source reference."],
 ["  note            Caveats; the word 'timing' triggers a small flag on the chart."],
 ["  severity        Bottleneck table only: 1 = most severe shading … 4 = least; blank = no shading."],
 ["  display_order   Order of rows within a chart."],
]
for line in info: notes.append(line)
notes["A1"].font = Font(name="Arial", bold=True, size=12)
for row in notes.iter_rows(min_row=2):
    for cell in row: cell.font = Font(name="Arial", size=10)
notes.column_dimensions["A"].width = 110
wb.save("data/city_comparison.xlsx")
print(f"OK: {len(rows)} rows")
